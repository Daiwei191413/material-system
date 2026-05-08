import pandas as pd
from openpyxl import load_workbook

new_bom = "/root/.hermes/profiles/development-assistant/cache/documents/doc_aea1134c47a4_器件BOM库V1.0 20260419.xlsx"

wb = load_workbook(new_bom, read_only=True)
print("=== 新 BOM 库 Sheets ===")
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"  - {sn}: {ws.max_row} 行 × {ws.max_column} 列")
wb.close()

xl = pd.ExcelFile(new_bom)
for sn in xl.sheet_names:
    df = pd.read_excel(new_bom, sheet_name=sn, header=None, nrows=6)
    print(f"\n=== Sheet: {sn} (前6行) ===")
    print(df.to_string(max_cols=20, max_colwidth=18))

# 对第一个数据 sheet 做完整读取
first = xl.sheet_names[0]
df_full = pd.read_excel(new_bom, sheet_name=first, header=0)
print(f"\n=== Sheet: {first} 列名 ===")
print(list(df_full.columns))
print(f"\n=== 总行数: {len(df_full)} ===")
print(f"\n=== 前3行示例 ===")
print(df_full.head(3).to_string(max_cols=20, max_colwidth=20))
